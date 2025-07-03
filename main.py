import streamlit as st
import pandas as pd
import plotly.express as px
from io import BytesIO
from openpyxl import load_workbook
from datetime import datetime

# ---------------------
# FUNCIONES AUXILIARES
# ---------------------
def identificar_empresa(wb):
    for hoja in wb.sheetnames:
        for celda in ["B2", "C2", "D2"]:
            valor = wb[hoja][celda].value
            if valor and "factura de crédito fiscal" in str(valor).lower():
                return "humano", hoja
    return "yunen", "Sheet1"

def extraer_datos_factura_from_workbook(wb):
    empresa, hoja_base = identificar_empresa(wb)
    hojas = wb.sheetnames

    datos = {
        "Cliente": "",
        "RNC Cliente": "",
        "Número Factura": "",
        "Fecha Emisión": "",
        "Dirección": "",
        "Teléfono": "",
        "Plan": "",
        "Monto": "",
        "NCF": "",
        "Vigencia": "",
        "Aseguradora": empresa.capitalize()
    }

    if empresa == "humano":
        hoja_uso = hoja_base if hoja_base in hojas else "Sheet2"
        sh = wb[hoja_uso]

        datos["Cliente"] = sh["B9"].value or ""
        datos["Dirección"] = sh["B9"].value or ""
        datos["RNC Cliente"] = (sh["B10"].value or sh["B11"].value or "")
        datos["NCF"] = sh["E7"].value or sh["D7"].value or ""
        datos["Monto"] = sh["E20"].value or ""
        datos["Fecha Emisión"] = sh["B8"].value or ""
        datos["Número Factura"] = f"{sh['A17'].value or ''} {sh['A18'].value or ''} {sh['A19'].value or ''}".strip()
        datos["Vigencia"] = f"{sh['B16'].value or ''} {sh['B17'].value or ''} {sh['B18'].value or ''}".strip()

        hoja_plan = None
        for hoja in wb.sheetnames:
            if str(wb[hoja]["A1"].value or "").strip().lower() == "detalle de facturación":
                hoja_plan = hoja
                break
        if hoja_plan:
            datos["Plan"] = f"{wb[hoja_plan]['A8'].value or ''} {wb[hoja_plan]['A9'].value or ''}".strip()

    else:
        sh = wb[hoja_base]
        datos["NCF"] = sh["B8"].value or ""
        datos["Fecha Emisión"] = sh["A9"].value or ""
        datos["Número Factura"] = sh["E10"].value or ""
        datos["RNC Cliente"] = sh["A11"].value or ""
        datos["Monto"] = wb["Sheet2"]["H24"].value or ""
        datos["Dirección"] = sh["B13"].value or ""
        datos["Teléfono"] = sh["B16"].value or ""
        datos["Plan"] = sh["A19"].value or ""
        datos["Cliente"] = sh["B12"].value or ""
        datos["Vigencia"] = f"{sh['D13'].value or ''} {sh['E13'].value or ''} {sh['F13'].value or ''}".strip()

    return datos

def procesar_archivos_cargados(files):
    consolidado = []
    for file in files:
        wb = load_workbook(file, data_only=True)
        datos = extraer_datos_factura_from_workbook(wb)
        consolidado.append(datos)
    return pd.DataFrame(consolidado)

def clasificar_vigencia(vigencia):
    try:
        partes = vigencia.split()
        if len(partes) >= 2:
            inicio = datetime.strptime(partes[0], "%d/%m/%Y")
            fin = datetime.strptime(partes[-1], "%d/%m/%Y")
            dias = (fin - inicio).days
            if dias < 45:
                return "Mensual"
            elif dias < 135:
                return "Trimestral"
            elif dias < 270:
                return "Semestral"
            else:
                return "Anual"
    except:
        return "Desconocido"

# ---------------------
# INICIO DE LA APP
# ---------------------
st.set_page_config(page_title="Resumen de Facturas", layout="wide")

menu = st.sidebar.selectbox("Navegación", ["📊 Resumen General", "👤 Análisis por Cliente", "📂 Cargar Archivos"])

# ---------------------
# Cargar archivo base
# ---------------------
@st.cache_data
def cargar_datos():
    df = pd.read_excel("datos_facturas.xlsx")
    df["Tipo Vigencia"] = df["Vigencia"].apply(clasificar_vigencia)
    return df

df = cargar_datos()

# ---------------------
# PÁGINA 1 - RESUMEN GENERAL
# ---------------------
if menu == "📊 Resumen General":
    st.title("📊 Resumen General de Facturas")

    col1, col2 = st.columns(2)

    with col1:
        st.subheader("🧾 Cantidad de Clientes por Aseguradora")
        fig1 = px.histogram(df, x="Aseguradora", color="Aseguradora", title="Clientes por Aseguradora")
        st.plotly_chart(fig1, use_container_width=True)

    with col2:
        st.subheader("📋 Cantidad de Planes por Aseguradora")
        fig2 = px.histogram(df, x="Aseguradora", color="Plan", title="Planes por Aseguradora", barmode="group")
        st.plotly_chart(fig2, use_container_width=True)

    st.subheader("📅 Distribución de Tipos de Vigencia")
    fig3 = px.pie(df, names="Tipo Vigencia", title="Distribución General de Vigencias")
    st.plotly_chart(fig3, use_container_width=True)

    st.subheader("📄 Tabla Resumen")
    st.dataframe(df[["Cliente", "Plan", "Vigencia", "Tipo Vigencia", "Aseguradora"]])

# ---------------------
# PÁGINA 2 - ANÁLISIS POR CLIENTE
# ---------------------
elif menu == "👤 Análisis por Cliente":
    st.title("👤 Reporte de Cliente")

    clientes = df["Cliente"].dropna().unique()
    cliente_seleccionado = st.selectbox("Selecciona un cliente:", sorted(clientes))

    df_cliente = df[df["Cliente"] == cliente_seleccionado]

    st.subheader("📄 Detalles del Cliente")
    st.dataframe(df_cliente)

    if not df_cliente.empty:
        col1, col2 = st.columns(2)
        with col1:
            fig = px.bar(
                df_cliente,
                x="Número Factura",
                y="Monto",
                text="Monto",
                title="Montos facturados",
                labels={"Monto": "RD$"},
            )
            fig.update_traces(texttemplate="%{text:.2f}", textposition="outside")
            fig.update_layout(uniformtext_minsize=8, uniformtext_mode="hide")
            st.plotly_chart(fig)

        with col2:
            st.metric("Cantidad de Facturas", len(df_cliente))
            st.metric("Total Facturado", f"RD${df_cliente['Monto'].sum():,.2f}")

        st.subheader("📆 Vigencias del Cliente")
        st.dataframe(df_cliente[["Plan", "Vigencia", "Tipo Vigencia"]])

        buffer = BytesIO()
        df_cliente.to_excel(buffer, index=False, engine="openpyxl")
        buffer.seek(0)

        st.download_button(
            label="📥 Descargar Excel del Cliente",
            data=buffer,
            file_name=f"{cliente_seleccionado}_datos.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

# ---------------------
# PÁGINA 3 - CARGA DE ARCHIVOS
# ---------------------
elif menu == "📂 Cargar Archivos":
    st.title("📂 Cargar Nuevos Archivos de Facturas")

    archivos = st.file_uploader("Selecciona uno o varios archivos Excel", type=["xlsx"], accept_multiple_files=True)

    if archivos:
        df_nuevo = procesar_archivos_cargados(archivos)
        df_nuevo["Tipo Vigencia"] = df_nuevo["Vigencia"].apply(clasificar_vigencia)
        st.success("Datos extraídos correctamente")

        st.subheader("📄 Vista previa de los datos")
        st.dataframe(df_nuevo)

        # Descargar nuevo Excel
        buffer = BytesIO()
        df_nuevo.to_excel(buffer, index=False, engine="openpyxl")
        buffer.seek(0)

        st.download_button(
            label="📥 Descargar Excel Consolidado",
            data=buffer,
            file_name="nuevas_facturas.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
